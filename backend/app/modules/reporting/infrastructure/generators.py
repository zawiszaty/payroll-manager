import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.modules.reporting.domain.entities import Report
from app.modules.reporting.domain.generators import IReportGenerator


class PDFReportGenerator(IReportGenerator):
    def __init__(self, output_dir: str = "/tmp/reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def supports_format(self, format_type: str) -> bool:
        return format_type.lower() == "pdf"

    async def generate(self, report: Report, data: dict[str, Any]) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"report_{report.id}_{timestamp}.pdf"
        file_path = self.output_dir / filename

        doc = SimpleDocTemplate(str(file_path), pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        title = Paragraph(f"<b>{report.name}</b>", styles["Title"])
        story.append(title)
        story.append(Spacer(1, 12))

        info_text = f"""
        <b>Report Type:</b> {report.report_type.value}<br/>
        <b>Generated:</b> {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}<br/>
        <b>Parameters:</b><br/>
        """

        if report.parameters.employee_id:
            info_text += f"- Employee ID: {report.parameters.employee_id}<br/>"
        if report.parameters.department:
            info_text += f"- Department: {report.parameters.department}<br/>"
        if report.parameters.start_date and report.parameters.end_date:
            info_text += (
                f"- Period: {report.parameters.start_date} to {report.parameters.end_date}<br/>"
            )

        info = Paragraph(info_text, styles["Normal"])
        story.append(info)
        story.append(Spacer(1, 20))

        if "rows" in data and data["rows"]:
            headers = data.get("headers", [])

            # Only prepend headers if they are non-empty
            if headers:
                table_data = [headers] + data["rows"]
                # Apply header-specific styling
                table_style = TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 14),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            else:
                # No headers, just use rows
                table_data = data["rows"]
                # Apply only body styling (no row 0 header styling)
                table_style = TableStyle(
                    [
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("BACKGROUND", (0, 0), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )

            table = Table(table_data)
            table.setStyle(table_style)
            story.append(table)

        doc.build(story)
        return file_path


class CSVReportGenerator(IReportGenerator):
    def __init__(self, output_dir: str = "/tmp/reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def supports_format(self, format_type: str) -> bool:
        return format_type.lower() == "csv"

    async def generate(self, report: Report, data: dict[str, Any]) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"report_{report.id}_{timestamp}.csv"
        file_path = self.output_dir / filename

        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            if "rows" in data and data["rows"]:
                headers = data.get("headers", [])
                writer = csv.writer(csvfile)

                writer.writerow([f"# {report.name}"])
                writer.writerow([f"# Report Type: {report.report_type.value}"])
                writer.writerow([f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                writer.writerow([])

                if headers:
                    writer.writerow(headers)
                writer.writerows(data["rows"])

        return file_path


class XLSXReportGenerator(IReportGenerator):
    def __init__(self, output_dir: str = "/tmp/reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def supports_format(self, format_type: str) -> bool:
        return format_type.lower() == "xlsx"

    async def generate(self, report: Report, data: dict[str, Any]) -> Path:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError(
                "openpyxl is required for XLSX generation. Install with: pip install openpyxl"
            )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"report_{report.id}_{timestamp}.xlsx"
        file_path = self.output_dir / filename

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        ws.append([report.name])
        ws["A1"].font = Font(bold=True, size=14)

        ws.append([f"Report Type: {report.report_type.value}"])
        ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([])

        if "rows" in data and data["rows"]:
            headers = data.get("headers", [])
            if headers:
                ws.append(headers)
                header_row = ws.max_row
                for cell in ws[header_row]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                    )

            for row in data["rows"]:
                ws.append(row)

        wb.save(str(file_path))
        return file_path


class JSONReportGenerator(IReportGenerator):
    def __init__(self, output_dir: str = "/tmp/reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def supports_format(self, format_type: str) -> bool:
        return format_type.lower() == "json"

    async def generate(self, report: Report, data: dict[str, Any]) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"report_{report.id}_{timestamp}.json"
        file_path = self.output_dir / filename

        report_data = {
            "metadata": {
                "name": report.name,
                "type": report.report_type.value,
                "generated_at": datetime.now().isoformat(),
                "parameters": {
                    "employee_id": report.parameters.employee_id,
                    "department": report.parameters.department,
                    "start_date": report.parameters.start_date,
                    "end_date": report.parameters.end_date,
                    "additional_filters": report.parameters.additional_filters,
                },
            },
            "data": data,
        }

        with open(file_path, "w", encoding="utf-8") as jsonfile:
            json.dump(report_data, jsonfile, indent=2, ensure_ascii=False)

        return file_path


class ReportGeneratorFactory:
    def __init__(self, output_dir: str = "/tmp/reports"):
        self.generators: list[IReportGenerator] = [
            PDFReportGenerator(output_dir),
            CSVReportGenerator(output_dir),
            XLSXReportGenerator(output_dir),
            JSONReportGenerator(output_dir),
        ]

    def get_generator(self, format_type: str) -> IReportGenerator:
        for generator in self.generators:
            if generator.supports_format(format_type):
                return generator
        raise ValueError(f"No generator found for format: {format_type}")
